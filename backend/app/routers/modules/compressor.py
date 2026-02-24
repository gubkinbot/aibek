"""Роутер модуля «Компрессорные станции».

REST API + WebSocket для управления станциями, тегами, аварийными правилами,
просмотра реалтайм-данных и исторических графиков.
"""

from __future__ import annotations

import json
import math
import uuid
from datetime import datetime, timezone
from io import BytesIO

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
    WebSocket,
    WebSocketDisconnect,
    status,
)
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import get_current_active_user, require_permission
from app.models.compressor import (
    CompressorAlarmEvent,
    CompressorAlarmRule,
    CompressorAnomalyEvent,
    CompressorAnomalyRule,
    CompressorComputedTag,
    CompressorStation,
    CompressorTag,
    CompressorTagValue,
)
from app.models.user import User
from app.schemas.compressor import (
    AlarmEventListResponse,
    AlarmEventResponse,
    AlarmRuleCreate,
    AlarmRuleResponse,
    AlarmRuleUpdate,
    AnomalyEventListResponse,
    AnomalyEventResponse,
    AnomalyRuleCreate,
    AnomalyRuleResponse,
    AnomalyRuleUpdate,
    ComputedTagCreate,
    ComputedTagResponse,
    ComputedTagUpdate,
    HistoryPoint,
    HistoryResponse,
    ImportResult,
    MessageResponse,
    RealtimeResponse,
    StationCreate,
    StationResponse,
    StationStatusResponse,
    StationUpdate,
    TagBulkCreate,
    TagCreate,
    TagResponse,
    TagUpdate,
)
from app.services.auth import decode_token
from app.services.redis import redis_client

router = APIRouter(prefix="/compressor", tags=["module-compressor"])


# ── Helpers ──────────────────────────────────────────────────────────────────


async def _get_station_or_404(
    code: str, db: AsyncSession,
) -> CompressorStation:
    """Возвращает станцию по коду или 404."""
    result = await db.execute(
        select(CompressorStation).where(CompressorStation.code == code),
    )
    station = result.scalar_one_or_none()
    if station is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"Станция «{code}» не найдена")
    return station


def _station_to_response(station: CompressorStation) -> StationResponse:
    """Преобразует ORM-объект станции в Pydantic-ответ с подсчётом тегов."""
    return StationResponse(
        id=station.id,
        name=station.name,
        code=station.code,
        opc_url=station.opc_url,
        opc_security_policy=station.opc_security_policy,
        opc_security_mode=station.opc_security_mode,
        opc_cert_path=station.opc_cert_path,
        opc_key_path=station.opc_key_path,
        polling_interval=station.polling_interval,
        realtime_interval=station.realtime_interval,
        is_active=station.is_active,
        description=station.description,
        created_at=station.created_at,
        updated_at=station.updated_at,
        tags_count=len(station.tags) if station.tags else 0,
    )


# ── Станции ──────────────────────────────────────────────────────────────────


@router.get("/stations", response_model=list[StationResponse])
async def list_stations(
    user: User = Depends(require_permission("compressor.view")),
    db: AsyncSession = Depends(get_db),
):
    """Список всех компрессорных станций."""
    result = await db.execute(
        select(CompressorStation).order_by(CompressorStation.name),
    )
    return [_station_to_response(s) for s in result.scalars().all()]


@router.get("/stations/{code}", response_model=StationResponse)
async def get_station(
    code: str,
    user: User = Depends(require_permission("compressor.view")),
    db: AsyncSession = Depends(get_db),
):
    """Получить станцию по коду."""
    station = await _get_station_or_404(code, db)
    return _station_to_response(station)


@router.get("/stations/{code}/status", response_model=StationStatusResponse)
async def get_station_status(
    code: str,
    user: User = Depends(require_permission("compressor.view")),
    db: AsyncSession = Depends(get_db),
):
    """Статус OPC-подключения станции (из Redis)."""
    station = await _get_station_or_404(code, db)
    raw = await redis_client.get(f"opc:status:{station.code}")
    if raw:
        data = json.loads(raw)
        return StationStatusResponse(
            code=station.code,
            connected=data.get("connected", False),
            last_seen=data.get("last_seen"),
            error=data.get("error"),
            tags_count=data.get("tags_count", 0),
        )
    return StationStatusResponse(code=station.code)


@router.post("/stations", response_model=StationResponse, status_code=201)
async def create_station(
    body: StationCreate,
    user: User = Depends(require_permission("compressor.admin")),
    db: AsyncSession = Depends(get_db),
):
    """Создать новую компрессорную станцию."""
    existing = await db.execute(
        select(CompressorStation).where(CompressorStation.code == body.code),
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status.HTTP_409_CONFLICT, f"Код «{body.code}» уже занят")

    station = CompressorStation(**body.model_dump())
    db.add(station)
    await db.commit()
    await db.refresh(station)
    return _station_to_response(station)


@router.patch("/stations/{code}", response_model=StationResponse)
async def update_station(
    code: str,
    body: StationUpdate,
    user: User = Depends(require_permission("compressor.admin")),
    db: AsyncSession = Depends(get_db),
):
    """Обновить конфигурацию станции."""
    station = await _get_station_or_404(code, db)
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(station, field, value)
    await db.commit()
    await db.refresh(station)
    return _station_to_response(station)


@router.delete("/stations/{code}", response_model=MessageResponse)
async def delete_station(
    code: str,
    user: User = Depends(require_permission("compressor.admin")),
    db: AsyncSession = Depends(get_db),
):
    """Удалить станцию и все связанные данные."""
    station = await _get_station_or_404(code, db)
    await db.delete(station)
    await db.commit()
    return MessageResponse(message=f"Станция «{code}» удалена")


# ── Сертификаты OPC UA ───────────────────────────────────────────────────────


@router.post("/stations/{code}/generate-cert", response_model=MessageResponse)
async def generate_station_cert(
    code: str,
    user: User = Depends(require_permission("compressor.admin")),
    db: AsyncSession = Depends(get_db),
):
    """Сгенерировать самоподписанный OPC UA сертификат для станции.

    После генерации файл .der нужно скачать (GET .../download-cert)
    и добавить в Trusted Clients на Kepware-сервере.
    """
    from app.collector.certs import generate_certs

    station = await _get_station_or_404(code, db)
    cert_path, key_path = generate_certs(station.code, station.name)

    # Сохраняем пути в БД
    station.opc_cert_path = cert_path
    station.opc_key_path = key_path
    await db.commit()

    return MessageResponse(message=f"Сертификат сгенерирован. Скачайте .der файл и добавьте в Trusted Clients на Kepware.")


@router.get("/stations/{code}/download-cert")
async def download_station_cert(
    code: str,
    user: User = Depends(require_permission("compressor.admin")),
    db: AsyncSession = Depends(get_db),
):
    """Скачать .der сертификат станции для переноса в Kepware Trusted Clients."""
    from pathlib import Path

    from fastapi.responses import FileResponse

    from app.collector.certs import get_cert_paths

    station = await _get_station_or_404(code, db)
    cert_path, _ = get_cert_paths(station.code)

    if not Path(cert_path).exists():
        raise HTTPException(status_code=404, detail="Сертификат не найден. Сначала сгенерируйте его.")

    return FileResponse(
        path=str(cert_path),
        media_type="application/x-x509-ca-cert",
        filename=f"{station.code}_client_cert.der",
    )


# ── Теги ─────────────────────────────────────────────────────────────────────


@router.get("/stations/{code}/tags", response_model=list[TagResponse])
async def list_tags(
    code: str,
    user: User = Depends(require_permission("compressor.view")),
    db: AsyncSession = Depends(get_db),
):
    """Список тегов станции."""
    station = await _get_station_or_404(code, db)
    result = await db.execute(
        select(CompressorTag)
        .where(CompressorTag.station_id == station.id)
        .order_by(CompressorTag.sort_order, CompressorTag.name),
    )
    return [TagResponse.model_validate(t) for t in result.scalars().all()]


@router.post("/stations/{code}/tags", response_model=TagResponse, status_code=201)
async def create_tag(
    code: str,
    body: TagCreate,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Создать новый тег для станции."""
    station = await _get_station_or_404(code, db)
    existing = await db.execute(
        select(CompressorTag).where(
            CompressorTag.station_id == station.id,
            CompressorTag.opc_path == body.opc_path,
        ),
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Тег с OPC-путём «{body.opc_path}» уже существует на этой станции",
        )
    tag = CompressorTag(station_id=station.id, **body.model_dump())
    db.add(tag)
    await db.commit()
    await db.refresh(tag)
    return TagResponse.model_validate(tag)


@router.post("/stations/{code}/tags/bulk", response_model=ImportResult)
async def create_tags_bulk(
    code: str,
    body: TagBulkCreate,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Массовое создание тегов для станции."""
    station = await _get_station_or_404(code, db)

    # Существующие OPC-пути на этой станции
    result = await db.execute(
        select(CompressorTag.opc_path).where(
            CompressorTag.station_id == station.id,
        ),
    )
    existing_paths = {row[0] for row in result.all()}

    created = 0
    skipped = 0
    errors: list[str] = []

    for tag_data in body.tags:
        if tag_data.opc_path in existing_paths:
            skipped += 1
            continue
        try:
            tag = CompressorTag(station_id=station.id, **tag_data.model_dump())
            db.add(tag)
            existing_paths.add(tag_data.opc_path)
            created += 1
        except Exception as e:
            errors.append(f"{tag_data.opc_path}: {e}")

    await db.commit()
    return ImportResult(created=created, skipped=skipped, errors=errors)


@router.patch("/tags/{tag_id}", response_model=TagResponse)
async def update_tag(
    tag_id: uuid.UUID,
    body: TagUpdate,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Обновить тег."""
    result = await db.execute(select(CompressorTag).where(CompressorTag.id == tag_id))
    tag = result.scalar_one_or_none()
    if tag is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Тег не найден")

    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(tag, field, value)
    await db.commit()
    await db.refresh(tag)
    return TagResponse.model_validate(tag)


@router.delete("/tags/{tag_id}", response_model=MessageResponse)
async def delete_tag(
    tag_id: uuid.UUID,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Удалить тег."""
    result = await db.execute(select(CompressorTag).where(CompressorTag.id == tag_id))
    tag = result.scalar_one_or_none()
    if tag is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Тег не найден")
    await db.delete(tag)
    await db.commit()
    return MessageResponse(message="Тег удалён")


# ── Шаблоны и импорт из Excel ────────────────────────────────────────────────


@router.get("/tags/template-excel")
async def download_tags_template(
    user: User = Depends(require_permission("compressor.manage")),
):
    """Скачать шаблон Excel для импорта тегов с примером данных."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "tags"

    # Заголовки
    headers = [
        ("opc_path", "OPC-путь тега в Kepware (обязательно)", 40),
        ("name", "Название для отображения (обязательно)", 30),
        ("unit", "Единица измерения", 12),
        ("data_type", "Тип данных (Float, Boolean, Int)", 15),
        ("category", "Категория для группировки", 22),
        ("sort_order", "Порядок сортировки (число)", 14),
        ("valid_min", "Мин. допустимое значение", 14),
        ("valid_max", "Макс. допустимое значение", 14),
        ("stale_timeout", "Таймаут залипания (сек)", 14),
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    desc_font = Font(italic=True, color="888888", size=10)

    for col_idx, (name, desc, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        # Описание во 2-й строке
        desc_cell = ws.cell(row=2, column=col_idx, value=desc)
        desc_cell.font = desc_font
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Примеры данных (строки 3-7)
    examples = [
        ("Channel1.Device1.TempIn", "Температура на входе", "°C", "Float", "Температура", 1, -50, 150, 120),
        ("Channel1.Device1.TempOut", "Температура на выходе", "°C", "Float", "Температура", 2, -50, 150, 120),
        ("Channel1.Device1.Pressure", "Давление газа", "кгс/см²", "Float", "Давление", 3, 0, 100, 60),
        ("Channel1.Device1.Vibration", "Вибрация подшипника", "мм/с", "Float", "Вибрация", 4, 0, 50, 30),
        ("Channel1.Device1.StatusWork", "Статус работы ГПА", None, "Boolean", "Статус", 10, None, None, None),
    ]

    example_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    for row_idx, row_data in enumerate(examples, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill

    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tags_template.xlsx"},
    )


@router.get("/history/template-excel")
async def download_history_template(
    user: User = Depends(require_permission("compressor.manage")),
):
    """Скачать шаблон Excel для импорта исторических данных."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "history"

    headers = [
        ("datetime", "Дата и время (обязательно)", 22),
        ("opc_path", "OPC-путь тега (обязательно)", 40),
        ("value", "Числовое значение (обязательно)", 16),
        ("health", "Качество: good/bad/stale/outlier", 16),
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    desc_font = Font(italic=True, color="888888", size=10)

    for col_idx, (name, desc, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        desc_cell = ws.cell(row=2, column=col_idx, value=desc)
        desc_cell.font = desc_font
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    examples = [
        (datetime(2025, 1, 15, 10, 0, 0), "Channel1.Device1.TempIn", 42.5, "good"),
        (datetime(2025, 1, 15, 10, 5, 0), "Channel1.Device1.TempIn", 43.1, "good"),
        (datetime(2025, 1, 15, 10, 0, 0), "Channel1.Device1.Pressure", 35.8, "good"),
        (datetime(2025, 1, 15, 10, 5, 0), "Channel1.Device1.Pressure", 999.0, "outlier"),
    ]

    example_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    for row_idx, row_data in enumerate(examples, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill

    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=history_template.xlsx"},
    )


@router.post("/stations/{code}/tags/import-excel", response_model=ImportResult)
async def import_tags_excel(
    code: str,
    file: UploadFile = File(...),
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Импорт определений тегов из Excel (.xlsx).

    Ожидаемые колонки: opc_path, name, unit, data_type, category,
    valid_min, valid_max, stale_timeout.
    """
    import openpyxl

    station = await _get_station_or_404(code, db)
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Файл пуст")

    # Первая строка — заголовки
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    if "opc_path" not in headers or "name" not in headers:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Файл должен содержать колонки opc_path и name",
        )

    col_map = {h: i for i, h in enumerate(headers)}

    # Существующие OPC-пути
    result = await db.execute(
        select(CompressorTag.opc_path).where(
            CompressorTag.station_id == station.id,
        ),
    )
    existing_paths = {row[0] for row in result.all()}

    created = 0
    skipped = 0
    errors: list[str] = []

    def _cell(row: tuple, col: str):
        idx = col_map.get(col)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None

    def _float(row: tuple, col: str) -> float | None:
        v = _cell(row, col)
        if v is None or v == "":
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    def _int(row: tuple, col: str) -> int | None:
        v = _cell(row, col)
        if v is None or v == "":
            return None
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return None

    for row_idx, row in enumerate(rows[1:], start=2):
        opc_path = _cell(row, "opc_path")
        name = _cell(row, "name")
        if not opc_path or not name:
            errors.append(f"Строка {row_idx}: пустой opc_path или name")
            continue
        if opc_path in existing_paths:
            skipped += 1
            continue
        try:
            tag = CompressorTag(
                station_id=station.id,
                opc_path=opc_path,
                name=name,
                unit=_cell(row, "unit"),
                data_type=_cell(row, "data_type"),
                category=_cell(row, "category"),
                valid_min=_float(row, "valid_min"),
                valid_max=_float(row, "valid_max"),
                stale_timeout=_int(row, "stale_timeout"),
            )
            db.add(tag)
            existing_paths.add(opc_path)
            created += 1
        except Exception as e:
            errors.append(f"Строка {row_idx}: {e}")

    await db.commit()
    wb.close()
    return ImportResult(created=created, skipped=skipped, errors=errors)


@router.post("/stations/{code}/history/import-excel", response_model=ImportResult)
async def import_history_excel(
    code: str,
    file: UploadFile = File(...),
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Импорт исторических значений из Excel (.xlsx).

    Ожидаемые колонки: datetime (или time), opc_path (или point), value, health (опционально).
    """
    import openpyxl

    station = await _get_station_or_404(code, db)
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Файл пуст")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    col_map = {h: i for i, h in enumerate(headers)}

    # Определяем имя колонки времени и OPC-пути
    time_col = "datetime" if "datetime" in col_map else "time"
    path_col = "opc_path" if "opc_path" in col_map else "point"

    if time_col not in col_map or path_col not in col_map or "value" not in col_map:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Файл должен содержать колонки datetime/time, opc_path/point, value",
        )

    # Маппинг opc_path → tag_id
    result = await db.execute(
        select(CompressorTag.opc_path, CompressorTag.id).where(
            CompressorTag.station_id == station.id,
        ),
    )
    path_to_id: dict[str, uuid.UUID] = {row[0]: row[1] for row in result.all()}

    imported = 0
    skipped = 0
    errors: list[str] = []
    batch: list[dict] = []
    batch_size = 1000

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            time_val = row[col_map[time_col]]
            opc_path = row[col_map[path_col]]
            value_raw = row[col_map["value"]]

            if not time_val or not opc_path:
                skipped += 1
                continue

            opc_path = str(opc_path).strip()
            tag_id = path_to_id.get(opc_path)
            if tag_id is None:
                skipped += 1
                continue

            # Парсинг времени
            if isinstance(time_val, datetime):
                ts = time_val.replace(tzinfo=timezone.utc) if time_val.tzinfo is None else time_val
            else:
                ts = datetime.fromisoformat(str(time_val).strip())
                if ts.tzinfo is None:
                    ts = ts.replace(tzinfo=timezone.utc)

            # Парсинг значения
            value = None
            if value_raw is not None:
                try:
                    value = float(value_raw)
                except (ValueError, TypeError):
                    pass

            # Quality из health колонки
            quality = "good"
            health_idx = col_map.get("health")
            if health_idx is not None and health_idx < len(row) and row[health_idx]:
                health_str = str(row[health_idx]).strip().lower()
                if health_str in ("bad", "stale", "outlier"):
                    quality = health_str

            batch.append({
                "time": ts,
                "tag_id": tag_id,
                "value": value,
                "quality": quality,
            })
            imported += 1

            if len(batch) >= batch_size:
                await db.execute(
                    CompressorTagValue.__table__.insert(), batch,
                )
                batch.clear()

        except Exception as e:
            errors.append(f"Строка {row_idx}: {e}")

    if batch:
        await db.execute(CompressorTagValue.__table__.insert(), batch)

    await db.commit()
    wb.close()
    return ImportResult(imported=imported, skipped=skipped, errors=errors)


# ── Вычисляемые теги ────────────────────────────────────────────────────────


@router.get(
    "/stations/{code}/computed-tags",
    response_model=list[ComputedTagResponse],
)
async def list_computed_tags(
    code: str,
    user: User = Depends(require_permission("compressor.view")),
    db: AsyncSession = Depends(get_db),
):
    """Список вычисляемых тегов станции."""
    station = await _get_station_or_404(code, db)
    result = await db.execute(
        select(CompressorComputedTag)
        .where(CompressorComputedTag.station_id == station.id)
        .order_by(CompressorComputedTag.sort_order, CompressorComputedTag.name),
    )
    return [ComputedTagResponse.model_validate(ct) for ct in result.scalars().all()]


@router.post(
    "/stations/{code}/computed-tags",
    response_model=ComputedTagResponse,
    status_code=201,
)
async def create_computed_tag(
    code: str,
    body: ComputedTagCreate,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Создать вычисляемый тег."""
    station = await _get_station_or_404(code, db)
    ct = CompressorComputedTag(station_id=station.id, **body.model_dump())
    db.add(ct)
    await db.commit()
    await db.refresh(ct)
    return ComputedTagResponse.model_validate(ct)


@router.patch("/computed-tags/{ct_id}", response_model=ComputedTagResponse)
async def update_computed_tag(
    ct_id: uuid.UUID,
    body: ComputedTagUpdate,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Обновить вычисляемый тег."""
    result = await db.execute(
        select(CompressorComputedTag).where(CompressorComputedTag.id == ct_id),
    )
    ct = result.scalar_one_or_none()
    if ct is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Вычисляемый тег не найден")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(ct, field, value)
    await db.commit()
    await db.refresh(ct)
    return ComputedTagResponse.model_validate(ct)


@router.delete("/computed-tags/{ct_id}", response_model=MessageResponse)
async def delete_computed_tag(
    ct_id: uuid.UUID,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Удалить вычисляемый тег."""
    result = await db.execute(
        select(CompressorComputedTag).where(CompressorComputedTag.id == ct_id),
    )
    ct = result.scalar_one_or_none()
    if ct is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Вычисляемый тег не найден")
    await db.delete(ct)
    await db.commit()
    return MessageResponse(message="Вычисляемый тег удалён")


# ── Realtime данные ──────────────────────────────────────────────────────────


@router.get("/stations/{code}/realtime", response_model=RealtimeResponse)
async def get_realtime(
    code: str,
    user: User = Depends(require_permission("compressor.view")),
    db: AsyncSession = Depends(get_db),
):
    """Текущий snapshot данных станции из Redis."""
    station = await _get_station_or_404(code, db)
    raw = await redis_client.get(f"opc:snapshot:{station.code}")
    if raw:
        return RealtimeResponse(**json.loads(raw))
    return RealtimeResponse(
        station=station.code,
        timestamp=None,
        connected=False,
        tags={},
        computed={},
    )


# ── Исторические данные ──────────────────────────────────────────────────────


@router.get("/stations/{code}/history", response_model=HistoryResponse)
async def get_history(
    code: str,
    tag_ids: str = Query(..., description="UUID тегов через запятую"),
    start: datetime = Query(..., alias="from", description="Начало периода (ISO 8601)"),
    end: datetime = Query(..., alias="to", description="Конец периода (ISO 8601)"),
    interval: str = Query("5 minutes", description="Интервал агрегации (например '5 minutes', '1 hour')"),
    user: User = Depends(require_permission("compressor.view")),
    db: AsyncSession = Depends(get_db),
):
    """Исторические данные с TimescaleDB time_bucket агрегацией."""
    station = await _get_station_or_404(code, db)

    # Парсинг tag_ids
    try:
        ids = [uuid.UUID(tid.strip()) for tid in tag_ids.split(",")]
    except ValueError:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Неверный формат tag_ids")

    # Проверяем, что теги принадлежат станции
    result = await db.execute(
        select(CompressorTag.id).where(
            CompressorTag.station_id == station.id,
            CompressorTag.id.in_(ids),
        ),
    )
    valid_ids = [row[0] for row in result.all()]
    if not valid_ids:
        return HistoryResponse(station=station.code, points=[])

    # TimescaleDB time_bucket запрос
    query = text("""
        SELECT
            time_bucket(:interval, time) AS bucket,
            tag_id,
            avg(value) AS avg,
            min(value) AS min,
            max(value) AS max
        FROM compressor_tag_values
        WHERE tag_id = ANY(:tag_ids)
          AND time >= :start
          AND time < :end
          AND quality = 'good'
        GROUP BY bucket, tag_id
        ORDER BY bucket
    """)

    result = await db.execute(query, {
        "interval": interval,
        "tag_ids": valid_ids,
        "start": start,
        "end": end,
    })

    points = [
        HistoryPoint(
            time=row.bucket,
            tag_id=row.tag_id,
            avg=row.avg,
            min=row.min,
            max=row.max,
        )
        for row in result.all()
    ]
    return HistoryResponse(station=station.code, points=points)


# ── Правила аварий ───────────────────────────────────────────────────────────


@router.get(
    "/stations/{code}/alarm-rules",
    response_model=list[AlarmRuleResponse],
)
async def list_alarm_rules(
    code: str,
    user: User = Depends(require_permission("compressor.view")),
    db: AsyncSession = Depends(get_db),
):
    """Список правил аварий для станции."""
    station = await _get_station_or_404(code, db)
    tag_ids = select(CompressorTag.id).where(
        CompressorTag.station_id == station.id,
    )
    result = await db.execute(
        select(CompressorAlarmRule)
        .where(CompressorAlarmRule.tag_id.in_(tag_ids))
        .order_by(CompressorAlarmRule.created_at.desc()),
    )
    return [AlarmRuleResponse.model_validate(r) for r in result.scalars().all()]


@router.post(
    "/stations/{code}/alarm-rules",
    response_model=AlarmRuleResponse,
    status_code=201,
)
async def create_alarm_rule(
    code: str,
    body: AlarmRuleCreate,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Создать правило аварии."""
    station = await _get_station_or_404(code, db)
    # Проверяем, что тег принадлежит станции
    result = await db.execute(
        select(CompressorTag).where(
            CompressorTag.id == body.tag_id,
            CompressorTag.station_id == station.id,
        ),
    )
    if result.scalar_one_or_none() is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Тег не найден на данной станции")

    rule = CompressorAlarmRule(**body.model_dump())
    db.add(rule)
    await db.commit()
    await db.refresh(rule)
    return AlarmRuleResponse.model_validate(rule)


@router.patch("/alarm-rules/{rule_id}", response_model=AlarmRuleResponse)
async def update_alarm_rule(
    rule_id: uuid.UUID,
    body: AlarmRuleUpdate,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Обновить правило аварии."""
    result = await db.execute(
        select(CompressorAlarmRule).where(CompressorAlarmRule.id == rule_id),
    )
    rule = result.scalar_one_or_none()
    if rule is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Правило не найдено")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(rule, field, value)
    await db.commit()
    await db.refresh(rule)
    return AlarmRuleResponse.model_validate(rule)


@router.delete("/alarm-rules/{rule_id}", response_model=MessageResponse)
async def delete_alarm_rule(
    rule_id: uuid.UUID,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Удалить правило аварии."""
    result = await db.execute(
        select(CompressorAlarmRule).where(CompressorAlarmRule.id == rule_id),
    )
    if result.scalar_one_or_none() is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Правило не найдено")
    await db.execute(
        delete(CompressorAlarmRule).where(CompressorAlarmRule.id == rule_id),
    )
    await db.commit()
    return MessageResponse(message="Правило аварии удалено")


# ── События аварий ───────────────────────────────────────────────────────────


@router.get("/stations/{code}/alarms", response_model=AlarmEventListResponse)
async def list_alarm_events(
    code: str,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    severity: str | None = None,
    acknowledged: bool | None = None,
    user: User = Depends(require_permission("compressor.view")),
    db: AsyncSession = Depends(get_db),
):
    """Журнал аварий станции с пагинацией и фильтрами."""
    station = await _get_station_or_404(code, db)

    base = select(CompressorAlarmEvent).where(
        CompressorAlarmEvent.station_id == station.id,
    )
    count_q = select(func.count()).select_from(CompressorAlarmEvent).where(
        CompressorAlarmEvent.station_id == station.id,
    )

    if severity:
        base = base.where(CompressorAlarmEvent.severity == severity)
        count_q = count_q.where(CompressorAlarmEvent.severity == severity)
    if acknowledged is not None:
        base = base.where(CompressorAlarmEvent.acknowledged == acknowledged)
        count_q = count_q.where(CompressorAlarmEvent.acknowledged == acknowledged)

    total = (await db.execute(count_q)).scalar() or 0
    pages = max(1, math.ceil(total / per_page))
    offset = (page - 1) * per_page

    result = await db.execute(
        base.order_by(CompressorAlarmEvent.time.desc())
        .offset(offset)
        .limit(per_page),
    )
    items = [AlarmEventResponse.model_validate(e) for e in result.scalars().all()]
    return AlarmEventListResponse(
        items=items, total=total, page=page, per_page=per_page, pages=pages,
    )


@router.post("/alarms/acknowledge", response_model=MessageResponse)
async def acknowledge_alarm(
    time: datetime = Query(...),
    station_id: uuid.UUID = Query(...),
    user: User = Depends(require_permission("compressor.edit")),
    db: AsyncSession = Depends(get_db),
):
    """Квитирование аварийного события."""
    result = await db.execute(
        select(CompressorAlarmEvent).where(
            CompressorAlarmEvent.time == time,
            CompressorAlarmEvent.station_id == station_id,
        ),
    )
    event = result.scalar_one_or_none()
    if event is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Событие не найдено")
    event.acknowledged = True
    event.acknowledged_by = user.id
    event.acknowledged_at = datetime.now(timezone.utc)
    await db.commit()
    return MessageResponse(message="Авария квитирована")


# ── Правила аномалий ─────────────────────────────────────────────────────────


@router.get(
    "/stations/{code}/anomaly-rules",
    response_model=list[AnomalyRuleResponse],
)
async def list_anomaly_rules(
    code: str,
    user: User = Depends(require_permission("compressor.view")),
    db: AsyncSession = Depends(get_db),
):
    """Список правил аномалий для станции."""
    station = await _get_station_or_404(code, db)
    tag_ids = select(CompressorTag.id).where(
        CompressorTag.station_id == station.id,
    )
    result = await db.execute(
        select(CompressorAnomalyRule)
        .where(CompressorAnomalyRule.tag_id.in_(tag_ids))
        .order_by(CompressorAnomalyRule.created_at.desc()),
    )
    return [AnomalyRuleResponse.model_validate(r) for r in result.scalars().all()]


@router.post(
    "/stations/{code}/anomaly-rules",
    response_model=AnomalyRuleResponse,
    status_code=201,
)
async def create_anomaly_rule(
    code: str,
    body: AnomalyRuleCreate,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Создать правило аномалии."""
    station = await _get_station_or_404(code, db)
    result = await db.execute(
        select(CompressorTag).where(
            CompressorTag.id == body.tag_id,
            CompressorTag.station_id == station.id,
        ),
    )
    if result.scalar_one_or_none() is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Тег не найден на данной станции")

    rule = CompressorAnomalyRule(**body.model_dump())
    db.add(rule)
    await db.commit()
    await db.refresh(rule)
    return AnomalyRuleResponse.model_validate(rule)


@router.patch("/anomaly-rules/{rule_id}", response_model=AnomalyRuleResponse)
async def update_anomaly_rule(
    rule_id: uuid.UUID,
    body: AnomalyRuleUpdate,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Обновить правило аномалии."""
    result = await db.execute(
        select(CompressorAnomalyRule).where(CompressorAnomalyRule.id == rule_id),
    )
    rule = result.scalar_one_or_none()
    if rule is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Правило не найдено")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(rule, field, value)
    await db.commit()
    await db.refresh(rule)
    return AnomalyRuleResponse.model_validate(rule)


@router.delete("/anomaly-rules/{rule_id}", response_model=MessageResponse)
async def delete_anomaly_rule(
    rule_id: uuid.UUID,
    user: User = Depends(require_permission("compressor.manage")),
    db: AsyncSession = Depends(get_db),
):
    """Удалить правило аномалии."""
    result = await db.execute(
        select(CompressorAnomalyRule).where(CompressorAnomalyRule.id == rule_id),
    )
    if result.scalar_one_or_none() is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Правило не найдено")
    await db.execute(
        delete(CompressorAnomalyRule).where(CompressorAnomalyRule.id == rule_id),
    )
    await db.commit()
    return MessageResponse(message="Правило аномалии удалено")


# ── События аномалий ─────────────────────────────────────────────────────────


@router.get("/stations/{code}/anomalies", response_model=AnomalyEventListResponse)
async def list_anomaly_events(
    code: str,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    severity: str | None = None,
    detector_type: str | None = None,
    acknowledged: bool | None = None,
    user: User = Depends(require_permission("compressor.view")),
    db: AsyncSession = Depends(get_db),
):
    """Журнал аномалий станции с пагинацией и фильтрами."""
    station = await _get_station_or_404(code, db)

    base = select(CompressorAnomalyEvent).where(
        CompressorAnomalyEvent.station_id == station.id,
    )
    count_q = select(func.count()).select_from(CompressorAnomalyEvent).where(
        CompressorAnomalyEvent.station_id == station.id,
    )

    if severity:
        base = base.where(CompressorAnomalyEvent.severity == severity)
        count_q = count_q.where(CompressorAnomalyEvent.severity == severity)
    if detector_type:
        base = base.where(CompressorAnomalyEvent.detector_type == detector_type)
        count_q = count_q.where(CompressorAnomalyEvent.detector_type == detector_type)
    if acknowledged is not None:
        base = base.where(CompressorAnomalyEvent.acknowledged == acknowledged)
        count_q = count_q.where(CompressorAnomalyEvent.acknowledged == acknowledged)

    total = (await db.execute(count_q)).scalar() or 0
    pages = max(1, math.ceil(total / per_page))
    offset = (page - 1) * per_page

    result = await db.execute(
        base.order_by(CompressorAnomalyEvent.time.desc())
        .offset(offset)
        .limit(per_page),
    )
    items = [AnomalyEventResponse.model_validate(e) for e in result.scalars().all()]
    return AnomalyEventListResponse(
        items=items, total=total, page=page, per_page=per_page, pages=pages,
    )


@router.post("/anomalies/acknowledge", response_model=MessageResponse)
async def acknowledge_anomaly(
    time: datetime = Query(...),
    station_id: uuid.UUID = Query(...),
    user: User = Depends(require_permission("compressor.edit")),
    db: AsyncSession = Depends(get_db),
):
    """Квитирование события аномалии."""
    result = await db.execute(
        select(CompressorAnomalyEvent).where(
            CompressorAnomalyEvent.time == time,
            CompressorAnomalyEvent.station_id == station_id,
        ),
    )
    event = result.scalar_one_or_none()
    if event is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Событие не найдено")
    event.acknowledged = True
    event.acknowledged_by = user.id
    event.acknowledged_at = datetime.now(timezone.utc)
    await db.commit()
    return MessageResponse(message="Аномалия квитирована")


# ── WebSocket ────────────────────────────────────────────────────────────────


async def _ws_authenticate(token: str, db: AsyncSession) -> User | None:
    """Валидация JWT из query param для WebSocket."""
    payload = decode_token(token)
    if payload is None:
        return None
    user_id = payload.get("sub")
    if user_id is None:
        return None
    result = await db.execute(select(User).where(User.id == user_id))
    return result.scalar_one_or_none()


@router.websocket("/ws/{station_code}")
async def compressor_ws(
    ws: WebSocket,
    station_code: str,
    token: str = Query(...),
):
    """WebSocket для реалтайм-данных станции.

    Подписывается на Redis pub/sub каналы:
      - opc:realtime:{code}
      - opc:alarms:{code}
      - opc:anomalies:{code}
    И пересылает сообщения клиенту.
    """
    from app.database import async_session

    async with async_session() as db:
        user = await _ws_authenticate(token, db)
        if user is None or not user.is_active:
            await ws.close(code=4001, reason="Unauthorized")
            return

        # Проверяем станцию
        result = await db.execute(
            select(CompressorStation).where(CompressorStation.code == station_code),
        )
        station = result.scalar_one_or_none()
        if station is None:
            await ws.close(code=4004, reason="Station not found")
            return

    await ws.accept()

    # Подписка на Redis pub/sub каналы
    import redis.asyncio as aioredis
    from app.config import settings

    pubsub_conn = aioredis.from_url(settings.redis_url, decode_responses=True)
    pubsub = pubsub_conn.pubsub()
    channels = [
        f"opc:realtime:{station_code}",
        f"opc:alarms:{station_code}",
        f"opc:anomalies:{station_code}",
    ]
    await pubsub.subscribe(*channels)

    try:
        async for message in pubsub.listen():
            if message["type"] != "message":
                continue
            try:
                await ws.send_text(message["data"])
            except WebSocketDisconnect:
                break
    except WebSocketDisconnect:
        pass
    finally:
        await pubsub.unsubscribe(*channels)
        await pubsub.close()
        await pubsub_conn.aclose()
