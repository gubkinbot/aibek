import logging
import os
from datetime import datetime
from typing import Dict, List, Optional, Any
from opcua import Client, ua
from opcua.crypto import security_policies
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Путь для хранения сертификатов
CERT_DIR = os.path.join(os.path.dirname(__file__), "certificates")
CERT_FILE = os.path.join(CERT_DIR, "client_cert.der")
KEY_FILE = os.path.join(CERT_DIR, "client_key.pem")


def ensure_certificates():
    """Создание самоподписанных сертификатов если их нет"""
    os.makedirs(CERT_DIR, exist_ok=True)
    
    if os.path.exists(CERT_FILE) and os.path.exists(KEY_FILE):
        logger.info("Сертификаты найдены")
        return True
    
    try:
        from cryptography import x509
        from cryptography.x509.oid import NameOID
        from cryptography.hazmat.primitives import hashes, serialization
        from cryptography.hazmat.primitives.asymmetric import rsa
        from cryptography.hazmat.backends import default_backend
        import datetime as dt
        
        logger.info("Генерация новых сертификатов...")
        
        # Генерация ключа
        key = rsa.generate_private_key(
            public_exponent=65537,
            key_size=2048,
            backend=default_backend()
        )
        
        # Создание сертификата
        subject = issuer = x509.Name([
            x509.NameAttribute(NameOID.COUNTRY_NAME, "RU"),
            x509.NameAttribute(NameOID.ORGANIZATION_NAME, "OPC Dashboard"),
            x509.NameAttribute(NameOID.COMMON_NAME, "OPC UA Client"),
        ])
        
        cert = (
            x509.CertificateBuilder()
            .subject_name(subject)
            .issuer_name(issuer)
            .public_key(key.public_key())
            .serial_number(x509.random_serial_number())
            .not_valid_before(dt.datetime.utcnow())
            .not_valid_after(dt.datetime.utcnow() + dt.timedelta(days=365))
            .add_extension(
                x509.SubjectAlternativeName([
                    x509.UniformResourceIdentifier("urn:opcua:client"),
                ]),
                critical=False,
            )
            .sign(key, hashes.SHA256(), default_backend())
        )
        
        # Сохранение сертификата (DER формат)
        with open(CERT_FILE, "wb") as f:
            f.write(cert.public_bytes(serialization.Encoding.DER))
        
        # Сохранение ключа (PEM формат)
        with open(KEY_FILE, "wb") as f:
            f.write(key.private_bytes(
                encoding=serialization.Encoding.PEM,
                format=serialization.PrivateFormat.TraditionalOpenSSL,
                encryption_algorithm=serialization.NoEncryption()
            ))
        
        logger.info(f"Сертификаты созданы в {CERT_DIR}")
        return True
        
    except Exception as e:
        logger.error(f"Ошибка создания сертификатов: {e}")
        return False


class OPCUAClient:
    """Клиент для подключения к OPC UA серверу (Kepware)"""
    
    def __init__(self, server_url: str, username: str = None, password: str = None):
        self.server_url = server_url
        self.username = username
        self.password = password
        self.client: Optional[Client] = None
        self.connected = False
        self._node_cache: Dict[str, Any] = {}
    
    def connect(self) -> bool:
        """Подключение к OPC UA серверу с шифрованием"""
        try:
            # Создаём сертификаты если нужно
            if not ensure_certificates():
                raise Exception("Не удалось создать сертификаты")
            
            self.client = Client(self.server_url)
            
            # Устанавливаем таймаут
            self.client.connect_timeout = 10000
            self.client.session_timeout = 30000
            
            # Настраиваем безопасное соединение
            # Basic128Rsa15 - поддерживается Kepware
            # SecurityMode: 2 = Sign, 3 = SignAndEncrypt
            self.client.set_security(
                security_policies.SecurityPolicyBasic128Rsa15,
                CERT_FILE,
                KEY_FILE,
                None,  # server_certificate - автоматически получим
                ua.MessageSecurityMode.Sign  # Сначала попробуем Sign (mode=2)
            )
            
            # Имя приложения для идентификации в Kepware
            self.client.application_uri = "urn:opcua:client"
            
            # Аутентификация по логину/паролю
            if self.username:
                self.client.set_user(self.username)
            if self.password:
                self.client.set_password(self.password)
            
            self.client.connect()
            self.connected = True
            logger.info(f"Подключено к OPC UA серверу: {self.server_url} (пользователь: {self.username or 'anonymous'}, режим: Sign)")
            return True
        except Exception as e:
            logger.error(f"Ошибка подключения к OPC UA: {e}")
            self.connected = False
            return False
    
    def disconnect(self):
        """Отключение от сервера"""
        if self.client and self.connected:
            try:
                self.client.disconnect()
                logger.info("Отключено от OPC UA сервера")
            except Exception as e:
                logger.error(f"Ошибка при отключении: {e}")
            finally:
                self.connected = False
                self._node_cache.clear()
    
    def _get_node(self, tag_name: str):
        """Получить узел по имени тега (с кэшированием)"""
        if tag_name not in self._node_cache:
            # Kepware формат NodeId: ns=2;s=Channel.Device.Tag
            node_id = f"ns=2;s={tag_name}"
            self._node_cache[tag_name] = self.client.get_node(node_id)
        return self._node_cache[tag_name]
    
    def read_tag(self, tag_name: str) -> Dict:
        """Чтение одного тега"""
        if not self.connected:
            return {
                'tag_name': tag_name,
                'value': None,
                'quality': 'Disconnected',
                'timestamp': datetime.now()
            }
        
        try:
            node = self._get_node(tag_name)
            data_value = node.get_data_value()
            
            value = data_value.Value.Value
            quality = 'Good' if data_value.StatusCode.is_good() else 'Bad'
            timestamp = data_value.SourceTimestamp or datetime.now()
            
            return {
                'tag_name': tag_name,
                'value': value,
                'quality': quality,
                'timestamp': timestamp
            }
        except Exception as e:
            logger.error(f"Ошибка чтения тега {tag_name}: {e}")
            return {
                'tag_name': tag_name,
                'value': None,
                'quality': 'Error',
                'timestamp': datetime.now()
            }
    
    def read_tags(self, tag_names: List[str]) -> List[Dict]:
        """Чтение нескольких тегов пакетно (одним запросом)"""
        if not self.connected:
            return [{
                'tag_name': tag,
                'value': None,
                'quality': 'Disconnected',
                'timestamp': datetime.now()
            } for tag in tag_names]
        
        results = []
        try:
            # Получаем все узлы
            nodes = [self._get_node(tag) for tag in tag_names]
            
            # Пакетное чтение всех значений за один вызов
            values = self.client.get_values(nodes)
            
            for tag_name, value in zip(tag_names, values):
                results.append({
                    'tag_name': tag_name,
                    'value': value,
                    'quality': 'Good',
                    'timestamp': datetime.now()
                })
        except Exception as e:
            logger.error(f"Ошибка пакетного чтения тегов: {type(e).__name__}: {e}")
            # Помечаем соединение как разорванное для переподключения
            self.connected = False
            self._node_cache.clear()
            # Возвращаем результаты с ошибками
            for tag_name in tag_names:
                results.append({
                    'tag_name': tag_name,
                    'value': None,
                    'quality': 'Error',
                    'timestamp': datetime.now()
                })
        
        return results
    
    def browse_tags(self, path: str = "") -> List[str]:
        """Просмотр доступных тегов на сервере (только пути)"""
        tags_info = self.browse_tags_detailed()
        return [t['path'] for t in tags_info]
    
    def _to_json_serializable(self, value):
        """
        Конвертация OPC UA значения в JSON-сериализуемый тип
        
        Args:
            value: Любое значение из OPC UA
            
        Returns:
            JSON-сериализуемое значение (str, int, float, bool, list, dict, None)
        """
        if value is None:
            return None
        
        # Примитивные типы - возвращаем как есть
        if isinstance(value, (bool, int, float, str)):
            return value
        
        # Байты
        if isinstance(value, bytes):
            try:
                return value.decode('utf-8')
            except:
                return value.hex()
        
        # Datetime
        if isinstance(value, datetime):
            return value.isoformat()
        
        # Списки и массивы
        if isinstance(value, (list, tuple)):
            return [self._to_json_serializable(item) for item in value]
        
        # Словари
        if isinstance(value, dict):
            return {k: self._to_json_serializable(v) for k, v in value.items()}
        
        # OPC UA специфичные типы - пробуем извлечь значение
        if hasattr(value, 'Value'):
            return self._to_json_serializable(value.Value)
        
        # LocalizedText
        if hasattr(value, 'Text'):
            return str(value.Text) if value.Text else ''
        
        # QualifiedName
        if hasattr(value, 'Name'):
            return str(value.Name) if value.Name else ''
        
        # NodeId
        if hasattr(value, 'to_string'):
            return value.to_string()
        
        # Для всех остальных сложных объектов - конвертируем в строку
        try:
            return str(value)
        except:
            return f"<{type(value).__name__}>"
    
    def _get_tag_full_info(self, node, tag_path: str, tag_name: str) -> Dict:
        """
        Получить полную информацию о теге включая все атрибуты OPC UA
        
        Args:
            node: OPC UA узел тега
            tag_path: Полный путь тега (Channel.Device.Tag)
            tag_name: Имя тега
            
        Returns:
            Dict с полной информацией о теге
        """
        tag_info = {
            'path': tag_path,
            'name': tag_name,
            'node_id': node.nodeid.to_string(),
            'description': '',
            'data_type': '',
            'value': None,
            'quality': '',
            'timestamp': '',
            'engineering_units': '',
            'eu_range_low': None,
            'eu_range_high': None,
            'access_level': '',
            'user_access_level': '',
            'is_readable': False,
            'is_writable': False,
        }
        
        # 1. Получаем описание из атрибута Description (стандарт OPC UA)
        try:
            desc = node.get_attribute(ua.AttributeIds.Description)
            if desc.Value.Value:
                tag_info['description'] = str(desc.Value.Value.Text) if hasattr(desc.Value.Value, 'Text') else str(desc.Value.Value)
        except:
            pass
        
        # 2. Получаем тип данных
        try:
            data_type = node.get_data_type_as_variant_type()
            tag_info['data_type'] = str(data_type).replace('VariantType.', '')
        except:
            pass
        
        # 3. Получаем текущее значение и качество
        try:
            data_value = node.get_data_value()
            raw_value = data_value.Value.Value
            tag_info['value'] = self._to_json_serializable(raw_value)
            tag_info['quality'] = 'Good' if data_value.StatusCode.is_good() else f'Bad ({data_value.StatusCode})'
            if data_value.SourceTimestamp:
                tag_info['timestamp'] = data_value.SourceTimestamp.isoformat()
        except:
            tag_info['quality'] = 'Error'
        
        # 4. Получаем уровни доступа
        try:
            access_level = node.get_attribute(ua.AttributeIds.AccessLevel).Value.Value
            user_access_level = node.get_attribute(ua.AttributeIds.UserAccessLevel).Value.Value
            
            # Расшифровка битов доступа
            access_flags = []
            if access_level & 0x01: 
                access_flags.append('Read')
                tag_info['is_readable'] = True
            if access_level & 0x02: 
                access_flags.append('Write')
                tag_info['is_writable'] = True
            if access_level & 0x04: access_flags.append('HistoryRead')
            if access_level & 0x08: access_flags.append('HistoryWrite')
            
            tag_info['access_level'] = ', '.join(access_flags) if access_flags else 'None'
            
            user_flags = []
            if user_access_level & 0x01: user_flags.append('Read')
            if user_access_level & 0x02: user_flags.append('Write')
            tag_info['user_access_level'] = ', '.join(user_flags) if user_flags else 'None'
        except:
            pass
        
        # 5. Ищем дополнительные атрибуты в дочерних узлах (Kepware специфика)
        try:
            for sub_child in node.get_children():
                sub_name = sub_child.get_browse_name().Name
                
                # Описание (Kepware)
                if sub_name == '_Description' and not tag_info['description']:
                    try:
                        desc_val = sub_child.get_value()
                        if desc_val:
                            tag_info['description'] = str(desc_val)
                    except:
                        pass
                
                # Единицы измерения (EngineeringUnits)
                elif sub_name in ('EngineeringUnits', '_EngineeringUnits', 'EU'):
                    try:
                        eu_val = sub_child.get_value()
                        if eu_val:
                            # EngineeringUnits может быть структурой с DisplayName
                            if hasattr(eu_val, 'DisplayName'):
                                tag_info['engineering_units'] = str(eu_val.DisplayName.Text)
                            else:
                                tag_info['engineering_units'] = str(eu_val)
                    except:
                        pass
                
                # Диапазон EURange
                elif sub_name in ('EURange', '_EURange'):
                    try:
                        range_val = sub_child.get_value()
                        if range_val:
                            if hasattr(range_val, 'Low'):
                                tag_info['eu_range_low'] = self._to_json_serializable(range_val.Low)
                                tag_info['eu_range_high'] = self._to_json_serializable(range_val.High)
                            elif isinstance(range_val, (list, tuple)) and len(range_val) >= 2:
                                tag_info['eu_range_low'] = self._to_json_serializable(range_val[0])
                                tag_info['eu_range_high'] = self._to_json_serializable(range_val[1])
                    except:
                        pass
        except:
            pass
        
        # 6. Пробуем получить EURange и EngineeringUnits через стандартные свойства OPC UA
        try:
            # Ищем свойства по BrowsePath
            refs = node.get_references(refs=ua.ObjectIds.HasProperty)
            for ref in refs:
                prop_node = self.client.get_node(ref.NodeId)
                prop_name = prop_node.get_browse_name().Name
                
                if prop_name == 'EngineeringUnits' and not tag_info['engineering_units']:
                    eu = prop_node.get_value()
                    if hasattr(eu, 'DisplayName'):
                        tag_info['engineering_units'] = str(eu.DisplayName.Text)
                    elif eu:
                        tag_info['engineering_units'] = str(eu)
                        
                elif prop_name == 'EURange' and tag_info['eu_range_low'] is None:
                    eu_range = prop_node.get_value()
                    if hasattr(eu_range, 'Low'):
                        tag_info['eu_range_low'] = self._to_json_serializable(eu_range.Low)
                        tag_info['eu_range_high'] = self._to_json_serializable(eu_range.High)
        except:
            pass
        
        return tag_info
    
    def browse_tags_detailed(self, include_values: bool = True) -> List[Dict]:
        """
        Просмотр всех доступных тегов с полной информацией.
        Использует рекурсивный обход с пакетным чтением атрибутов.
        
        Args:
            include_values: Читать ли текущие значения
            
        Returns:
            List[Dict]: Список тегов с полной информацией
        """
        if not self.connected:
            return []
        
        import time
        start_time = time.time()
        
        try:
            objects = self.client.get_objects_node()
            
            # Фаза 1: Рекурсивный обход дерева
            logger.info("Фаза 1: Обход дерева OPC UA...")
            t1 = time.time()
            
            tag_nodes = []  # [(node, path, name), ...]
            
            def browse_recursive(node, path="", depth=0):
                """Рекурсивный обход узлов"""
                if depth > 15:
                    return
                
                try:
                    children = node.get_children()
                except Exception as e:
                    return
                
                for child in children:
                    try:
                        name = child.get_browse_name().Name
                        if name.startswith('_'):
                            continue
                        
                        new_path = f"{path}.{name}" if path else name
                        node_class = child.get_node_class()
                        
                        # NodeClass.Variable = 2 (это теги)
                        if node_class == ua.NodeClass.Variable:
                            tag_nodes.append((child, new_path, name))
                        
                        # Рекурсивно обходим дочерние узлы
                        if node_class in (ua.NodeClass.Object, ua.NodeClass.Variable):
                            browse_recursive(child, new_path, depth + 1)
                            
                    except Exception as e:
                        continue
            
            browse_recursive(objects)
            
            t2 = time.time()
            logger.info(f"Фаза 1 завершена за {t2-t1:.2f}с. Найдено {len(tag_nodes)} тегов.")
            
            if not tag_nodes:
                return []
            
            # Фаза 2: Пакетное чтение атрибутов
            logger.info("Фаза 2: Пакетное чтение атрибутов...")
            
            nodes = [t[0] for t in tag_nodes]
            
            # Пакетное чтение значений
            values = []
            try:
                values = self.client.get_values(nodes)
            except Exception as e:
                logger.error(f"Ошибка пакетного чтения значений: {e}")
                values = [None] * len(nodes)
            
            t3 = time.time()
            logger.info(f"Фаза 2 завершена за {t3-t2:.2f}с.")
            
            # Фаза 3: Формируем результаты
            logger.info("Фаза 3: Формирование результатов...")
            
            dt_map = {
                1: 'Boolean', 2: 'SByte', 3: 'Byte', 4: 'Int16', 5: 'UInt16',
                6: 'Int32', 7: 'UInt32', 8: 'Int64', 9: 'UInt64',
                10: 'Float', 11: 'Double', 12: 'String', 13: 'DateTime',
                14: 'Guid', 15: 'ByteString'
            }
            
            tags = []
            for i, (node, path, name) in enumerate(tag_nodes):
                tag_info = {
                    'path': path,
                    'name': name,
                    'node_id': node.nodeid.to_string(),
                    'description': '',
                    'data_type': '',
                    'value': None,
                    'quality': 'Good',
                    'timestamp': '',
                    'engineering_units': '',
                    'eu_range_low': None,
                    'eu_range_high': None,
                    'access_level': '',
                    'is_readable': False,
                    'is_writable': False,
                }
                
                # Значение из пакетного чтения
                try:
                    if i < len(values) and values[i] is not None:
                        tag_info['value'] = self._to_json_serializable(values[i])
                except:
                    pass
                
                # Дополнительные атрибуты (читаем по одному - быстрее чем кажется)
                try:
                    # Тип данных
                    data_type = node.get_data_type_as_variant_type()
                    tag_info['data_type'] = str(data_type).replace('VariantType.', '')
                except:
                    pass
                
                try:
                    # Описание
                    desc = node.get_attribute(ua.AttributeIds.Description)
                    if desc.Value.Value:
                        tag_info['description'] = str(desc.Value.Value.Text) if hasattr(desc.Value.Value, 'Text') else str(desc.Value.Value)
                except:
                    pass
                
                try:
                    # Уровень доступа
                    access_level = node.get_attribute(ua.AttributeIds.AccessLevel).Value.Value
                    if access_level is not None:
                        flags = []
                        if access_level & 0x01:
                            flags.append('Read')
                            tag_info['is_readable'] = True
                        if access_level & 0x02:
                            flags.append('Write')
                            tag_info['is_writable'] = True
                        tag_info['access_level'] = ', '.join(flags) if flags else 'None'
                except:
                    pass
                
                tags.append(tag_info)
            
            total_time = time.time() - start_time
            logger.info(f"✅ Сканирование завершено за {total_time:.2f}с. Тегов: {len(tags)}")
            return tags
            
        except Exception as e:
            logger.error(f"Ошибка просмотра тегов: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def export_tags_to_excel(self, filename: str = "opc_tags_export.xlsx", tags: List[Dict] = None) -> str:
        """
        Экспорт списка тегов в Excel файл с форматированием
        
        Args:
            filename: Имя файла (по умолчанию opc_tags_export.xlsx)
            tags: Список тегов (если None, будет выполнено сканирование)
            
        Returns:
            str: Путь к созданному файлу
        """
        if tags is None:
            tags = self.browse_tags_detailed()
        
        if not tags:
            raise ValueError("Нет тегов для экспорта")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "OPC UA Tags"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_alignment = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Заголовки
        headers = [
            ("№", 5),
            ("Путь тега", 45),
            ("Имя", 20),
            ("Описание", 40),
            ("Тип данных", 12),
            ("Значение", 15),
            ("Качество", 12),
            ("Ед. изм.", 10),
            ("Min", 10),
            ("Max", 10),
            ("Доступ", 15),
            ("Запись", 8),
            ("Node ID", 35),
            ("Timestamp", 22),
        ]
        
        # Записываем заголовки
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Фиксируем заголовок
        ws.freeze_panes = 'A2'
        
        # Записываем данные
        for row_idx, tag in enumerate(tags, 2):
            # Безопасное получение значения для Excel
            value = tag.get('value')
            if value is not None:
                value_str = self._to_json_serializable(value)
                if isinstance(value_str, (list, dict)):
                    value_str = str(value_str)
            else:
                value_str = ''
            
            row_data = [
                row_idx - 1,  # №
                str(tag.get('path', '')),
                str(tag.get('name', '')),
                str(tag.get('description', '')),
                str(tag.get('data_type', '')),
                str(value_str) if value_str else '',
                str(tag.get('quality', '')),
                str(tag.get('engineering_units', '')),
                self._to_json_serializable(tag.get('eu_range_low')) if tag.get('eu_range_low') is not None else '',
                self._to_json_serializable(tag.get('eu_range_high')) if tag.get('eu_range_high') is not None else '',
                str(tag.get('access_level', '')),
                'Да' if tag.get('is_writable') else 'Нет',
                str(tag.get('node_id', '')),
                str(tag.get('timestamp', '')),
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # Подсветка качества
                if col == 7:  # Колонка "Качество"
                    if 'Good' in str(value):
                        cell.fill = good_fill
                    elif value:
                        cell.fill = bad_fill
        
        # Добавляем лист со статистикой
        ws_stats = wb.create_sheet(title="Статистика")
        
        stats = [
            ("Общее количество тегов", len(tags)),
            ("Тегов с описанием", sum(1 for t in tags if t.get('description'))),
            ("Тегов без описания", sum(1 for t in tags if not t.get('description'))),
            ("Тегов для чтения", sum(1 for t in tags if t.get('is_readable'))),
            ("Тегов для записи", sum(1 for t in tags if t.get('is_writable'))),
            ("Качество Good", sum(1 for t in tags if 'Good' in str(t.get('quality', '')))),
            ("Качество Bad/Error", sum(1 for t in tags if t.get('quality') and 'Good' not in str(t.get('quality', '')))),
            ("", ""),
            ("Типы данных:", ""),
        ]
        
        # Подсчёт по типам данных
        data_types = {}
        for tag in tags:
            dt = tag.get('data_type', 'Unknown')
            data_types[dt] = data_types.get(dt, 0) + 1
        
        for dt, count in sorted(data_types.items()):
            stats.append((f"  {dt}", count))
        
        for row_idx, (label, value) in enumerate(stats, 1):
            ws_stats.cell(row=row_idx, column=1, value=label).font = Font(bold=True) if not label.startswith("  ") else Font()
            ws_stats.cell(row=row_idx, column=2, value=value)
        
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 15
        
        # Добавляем лист с уникальными каналами/устройствами
        ws_channels = wb.create_sheet(title="Структура")
        
        channels = {}
        for tag in tags:
            path_parts = tag.get('path', '').split('.')
            if len(path_parts) >= 2:
                channel = path_parts[0]
                device = path_parts[1] if len(path_parts) > 1 else ''
                key = f"{channel}.{device}"
                if key not in channels:
                    channels[key] = {'channel': channel, 'device': device, 'tags': 0}
                channels[key]['tags'] += 1
        
        ws_channels.cell(row=1, column=1, value="Канал").font = header_font
        ws_channels.cell(row=1, column=1).fill = header_fill
        ws_channels.cell(row=1, column=2, value="Устройство").font = header_font
        ws_channels.cell(row=1, column=2).fill = header_fill
        ws_channels.cell(row=1, column=3, value="Кол-во тегов").font = header_font
        ws_channels.cell(row=1, column=3).fill = header_fill
        
        for row_idx, (key, info) in enumerate(sorted(channels.items()), 2):
            ws_channels.cell(row=row_idx, column=1, value=info['channel'])
            ws_channels.cell(row=row_idx, column=2, value=info['device'])
            ws_channels.cell(row=row_idx, column=3, value=info['tags'])
        
        ws_channels.column_dimensions['A'].width = 25
        ws_channels.column_dimensions['B'].width = 25
        ws_channels.column_dimensions['C'].width = 15
        
        # Сохраняем файл
        filepath = os.path.join(os.path.dirname(__file__), filename)
        wb.save(filepath)
        logger.info(f"Excel файл сохранён: {filepath}")
        
        return filepath
